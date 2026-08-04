# export.py — The budget leaving the building: a workbook and a board pack.
#
# The reason to export from this tool rather than from a spreadsheet is the
# Assumptions sheet: every driver arrives with `source_url`, `retrieved_at` and
# `locked_at` attached. The data already carries that provenance — the guards in
# drivers.py are what make it true rather than decorative — so carrying it into
# the board pack is nearly free and is the whole differentiator. Do not "tidy"
# those columns away.
#
# Shape discipline, deliberately mirroring budget.py:
#
#   * `scenario_pack` / `version_pack` are pure translators — a stored record in,
#     a flat presentation dict out. They take driver provenance and company
#     details as ARGUMENTS, so this module reads no dataset and imports nothing
#     but budget.py (itself a pure leaf). That is what makes tests/test_export.py
#     possible without touching disk.
#   * `board_pack_markdown` is pure text and works with no third-party package,
#     so the markdown export cannot be taken down by a missing dependency.
#   * openpyxl is imported INSIDE `workbook()`. An existing checkout that has not
#     re-run `pip install -r requirements.txt` then loses one route rather than
#     failing to import main.py and taking the whole app down.

from __future__ import annotations

from datetime import datetime, timezone

from . import budget

# The P&L metrics a summary reports, in reading order, with their labels.
SUMMARY_ROWS = (
    ("volume_tonnes", "Volume (t)"),
    ("revenue_eur", "Revenue (€)"),
    ("cogs_eur", "COGS (€)"),
    ("gross_margin_eur", "Gross margin (€)"),
    ("opex_eur", "Opex (€)"),
    ("ebitda_eur", "EBITDA (€)"),
    ("ebitda_margin_pct", "EBITDA margin (%)"),
)

BLOCK_LABELS = {
    "drivers": "Driver price",
    "volume": "Volume",
    "price": "Selling price",
    "opex": "Opex",
}

BLOCK_KEY_LABELS = {
    "drivers": "Driver",
    "volume": "Product line",
    "price": "Product line",
    "opex": "Cost centre / driver",
}


# --------------------------------------------------------------------------
# Packs — the one shape both exporters render
# --------------------------------------------------------------------------

def scenario_pack(scenario: dict, *, drivers_snapshot=None, locked_at=None,
                  company: str | None = None, narrative: str | None = None,
                  diff: dict | None = None, cash: dict | None = None) -> dict:
    """A stored scenario, flattened for export.

    `cash` is `cash.project_cashflow`'s result (or the tool's), passed in as an
    ARGUMENT for the same reason `drivers_snapshot` is: this module reads no
    dataset. None means the cash profile was not available — a company with no
    working-capital history and no stated terms — and the sheet says so rather
    than printing zeros.
    """
    return _pack(scenario, kind="scenario",
                 title=scenario.get("name") or "Budget scenario",
                 assumptions=scenario.get("assumptions"),
                 drivers_snapshot=drivers_snapshot, locked_at=locked_at,
                 company=company, narrative=narrative, diff=diff, cash=cash,
                 meta=[("Scenario", scenario.get("name")),
                       ("Baseline", scenario.get("baseline")),
                       # What the drivers were priced at before the assumptions
                       # applied. A board pack that omits it cannot be read: the
                       # same percentages on a forward curve and on locked
                       # values are two different budgets.
                       ("Driver basis", scenario.get("basis") or "locked"),
                       ("Active", "yes" if scenario.get("active") else "no"),
                       ("Updated", _stamp(scenario.get("updated_at")))])


def version_pack(version: dict, *, company: str | None = None,
                 narrative: str | None = None, diff: dict | None = None) -> dict:
    """A frozen budget version, flattened for export. Its driver provenance AND
    its cash profile come from the version's own snapshots — never from today's
    dataset, which is the point of a version: the days were measured on the day
    the board approved it."""
    label = version.get("label") or f"Version {version.get('version_no')}"
    return _pack(version, kind="version",
                 title=f"v{version.get('version_no')} · {label}",
                 assumptions=version.get("assumptions_snapshot"),
                 drivers_snapshot=version.get("drivers_snapshot"),
                 locked_at=version.get("locked_at"),
                 company=company, narrative=narrative, diff=diff,
                 cash=version.get("cash_snapshot"),
                 meta=[("Version", version.get("version_no")),
                       ("Label", label),
                       ("Status", version.get("status")),
                       ("From scenario", version.get("scenario_name")),
                       ("Baseline", version.get("baseline")),
                       ("Driver basis", version.get("basis") or "locked"),
                       ("Created", _stamp(version.get("created_at"))),
                       ("Created by", version.get("created_by")),
                       ("Submitted by", version.get("submitted_by")),
                       ("Approved by", version.get("approved_by")),
                       ("Approved", _stamp(version.get("approved_at")))])


def _pack(record: dict, *, kind: str, title: str, assumptions, drivers_snapshot,
          locked_at, company, narrative, diff, meta, cash=None) -> dict:
    return {
        "kind": kind,
        "title": title,
        "company": company or "",
        "note": record.get("note"),
        "status": record.get("status"),
        "assumptions": budget.normalise_assumptions(assumptions),
        "price_pass_through": float(record.get("price_pass_through") or 0.0),
        "opex_inflation_pct": float(record.get("opex_inflation_pct") or 0.0),
        "totals": record.get("totals") or {},
        "by_month": record.get("by_month") or [],
        "by_product_line": record.get("by_product_line") or [],
        "driver_impact_eur": record.get("driver_impact_eur") or {},
        "opex_bridge": record.get("opex_bridge") or {},
        "ebitda_bridge": record.get("ebitda_bridge") or {},
        "drivers": list(drivers_snapshot or []),
        "cash": cash or None,
        "locked_at": locked_at,
        "narrative": narrative,
        "diff": diff,
        "meta": [(str(k), v) for k, v in meta if v not in (None, "")],
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }


def filename(pack: dict, extension: str) -> str:
    """A filesystem-safe download name — `budget-v3-2027-plan.xlsx`."""
    stem = "-".join(str(pack.get("title") or pack.get("kind") or "budget").split())
    safe = "".join(c if (c.isalnum() or c in "-_.") else "-" for c in stem).strip("-")
    safe = "-".join(part for part in safe.split("-") if part)[:60] or "budget"
    return f"{safe}.{extension.lstrip('.')}"


# --------------------------------------------------------------------------
# The workbook
# --------------------------------------------------------------------------

def workbook(pack: dict) -> bytes:
    """Seven sheets: Summary · Monthly P&L · By product line · Cash flow ·
    Assumptions · Driver bridge · Version diff. Returns the .xlsx bytes.

    Raises RuntimeError — never an ImportError — when openpyxl is absent, so the
    route can answer with something a CFO can act on.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:                                  # pragma: no cover
        raise RuntimeError(
            "Excel export needs openpyxl, which is not installed. Run "
            "`.venv/bin/pip install -r requirements.txt` and try again. The "
            "board-pack markdown export works without it."
        ) from exc

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3352")          # columbus navy
    title_font = Font(bold=True, size=14)
    money = "#,##0"
    pct = "#,##0.00"

    wb = Workbook()

    def sheet(name, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = name
        return ws

    def header(ws, row_idx, labels):
        for col, label in enumerate(labels, start=1):
            cell = ws.cell(row=row_idx, column=col, value=label)
            cell.font, cell.fill = head_font, head_fill
            cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = ws.cell(row=row_idx + 1, column=1)

    def widths(ws, *sizes):
        for i, size in enumerate(sizes, start=1):
            ws.column_dimensions[get_column_letter(i)].width = size

    def write(ws, row_idx, values, formats=None):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            fmt = (formats or {}).get(col)
            if fmt and isinstance(value, (int, float)):
                cell.number_format = fmt
        return row_idx + 1

    # ---- Summary ----
    ws = sheet("Summary", first=True)
    widths(ws, 30, 22, 22, 18)
    ws["A1"] = pack.get("company") or "Budget"
    ws["A1"].font = title_font
    ws["A2"] = pack.get("title") or ""
    ws["A2"].font = Font(bold=True)
    r = 4
    for label, value in pack.get("meta") or []:
        r = write(ws, r, [label, value])
    if pack.get("note"):
        r = write(ws, r, ["Note", pack["note"]])
    r = write(ws, r, ["Price pass-through (τ)", pack.get("price_pass_through")])
    r = write(ws, r, ["Opex inflation default (%)", pack.get("opex_inflation_pct")])
    r = write(ws, r, ["Exported (UTC)", pack.get("generated_at")])
    r += 1
    header(ws, r, ["Metric", "Value"])
    r += 1
    totals = pack.get("totals") or {}
    for key, label in SUMMARY_ROWS:
        r = write(ws, r, [label, _num(totals.get(key))],
                  {2: pct if key.endswith("_pct") else money})
    if pack.get("narrative"):
        r += 1
        r = write(ws, r, ["Read", pack["narrative"]])

    # ---- Monthly P&L ----
    ws = sheet("Monthly P&L")
    widths(ws, 14, 16, 16, 16, 16, 16, 16)
    header(ws, 1, ["Month", "Volume (t)", "Revenue (€)", "COGS (€)",
                   "Gross margin (€)", "Opex (€)", "EBITDA (€)"])
    r = 2
    fmt = {2: money, 3: money, 4: money, 5: money, 6: money, 7: money}
    for row in pack.get("by_month") or []:
        r = write(ws, r, [row.get("month"), _num(row.get("volume_tonnes")),
                          _num(row.get("revenue_eur")), _num(row.get("cogs_eur")),
                          _num(row.get("gross_margin_eur")), _num(row.get("opex_eur")),
                          _num(row.get("ebitda_eur"))], fmt)

    # ---- By product line ----
    ws = sheet("By product line")
    widths(ws, 24, 16, 16, 16, 16, 16, 16, 16)
    header(ws, 1, ["Product line", "Volume (t)", "Revenue (€)", "COGS (€)",
                   "Gross margin (€)", "Opex (€)", "EBITDA (€)", "EBITDA margin (%)"])
    r = 2
    fmt = {2: money, 3: money, 4: money, 5: money, 6: money, 7: money, 8: pct}
    for row in pack.get("by_product_line") or []:
        r = write(ws, r, [row.get("product_line"), _num(row.get("volume_tonnes")),
                          _num(row.get("revenue_eur")), _num(row.get("cogs_eur")),
                          _num(row.get("gross_margin_eur")), _num(row.get("opex_eur")),
                          _num(row.get("ebitda_eur")), _num(row.get("ebitda_margin_pct"))],
                  fmt)

    # ---- Cash flow ----
    ws = sheet("Cash flow")
    widths(ws, 14, 16, 20, 14, 14, 18, 18)
    cash = pack.get("cash")
    if not cash:
        ws["A1"] = ("No cash profile is attached. It needs either monthly receivables, "
                    "inventory and payables balances in the working_capital dataset, or "
                    "collection, inventory and payment days stated on the Budget page.")
    else:
        totals = cash.get("totals") or {}
        cdays = cash.get("days") or {}
        r = 1
        for label, value, fmt in (
                ("Days sales outstanding (DSO)", cdays.get("dso_days"), pct),
                ("Days inventory outstanding (DIO)", cdays.get("dio_days"), pct),
                ("Days payable outstanding (DPO)", cdays.get("dpo_days"), pct),
                ("Cash conversion cycle (days)", cdays.get("cash_conversion_cycle_days"), pct),
                # Whether each day count was measured off the balance sheet or
                # decided by the CFO is the whole provenance of this sheet — the
                # same argument the Assumptions sheet makes for driver prices.
                ("Days basis", cdays.get("basis"), None),
                ("Cash tax rate (%)", cash.get("tax_rate_pct"), pct),
                ("Minimum cash tested (€)", cash.get("min_cash_eur"), money)):
            r = write(ws, r, [label, _num(value) if fmt else value],
                      {2: fmt} if fmt else None)
        trough = cash.get("trough") or {}
        r = write(ws, r, ["Lowest cash balance (€)", _num(trough.get("closing_cash_eur"))],
                  {2: money})
        r = write(ws, r, ["Lowest month", trough.get("month")])
        if cash.get("months_below_minimum"):
            r = write(ws, r, ["Months below the minimum",
                              ", ".join(cash["months_below_minimum"])])
        r = write(ws, r, ["EBITDA converted to cash (%)",
                          _num(totals.get("cash_conversion_pct"))], {2: pct})
        r += 1

        header(ws, r, ["Month", "EBITDA (€)", "Working capital (€)", "Cash tax (€)",
                       "Capex (€)", "Free cash flow (€)", "Closing cash (€)"])
        r += 1
        fmt = {2: money, 3: money, 4: money, 5: money, 6: money, 7: money}
        for row in cash.get("months") or []:
            r = write(ws, r, [row.get("month"), _num(row.get("ebitda_eur")),
                              # Sign-flipped to the effect ON CASH: absorbing
                              # working capital is a negative in a cash column,
                              # however the balance moved.
                              -_num(row.get("working_capital_change_eur")),
                              -_num(row.get("tax_eur")), -_num(row.get("capex_eur")),
                              _num(row.get("free_cash_flow_eur")),
                              _num(row.get("closing_cash_eur"))], fmt)
        r = write(ws, r, ["Total", _num(totals.get("ebitda_eur")),
                          -_num(totals.get("working_capital_change_eur")),
                          -_num(totals.get("tax_eur")), -_num(totals.get("capex_eur")),
                          _num(totals.get("free_cash_flow_eur")),
                          _num(totals.get("closing_cash_eur"))], fmt)

        capex = cash.get("capex") or {}
        if capex.get("projects"):
            r += 1
            r = write(ws, r, ["Capital plan"])
            header(ws, r, ["Month", "Project", "Category", "Status", "Amount (€)"])
            r += 1
            for project in capex["projects"]:
                r = write(ws, r, [project.get("month"), project.get("project"),
                                  project.get("category"), project.get("status"),
                                  _num(project.get("amount_eur"))], {5: money})

        # The limitations travel with the figures rather than being left to the
        # covering email: this is free cash flow before financing, and a board
        # reading a closing balance has to know that.
        for caveat in cash.get("caveats") or []:
            r += 1
            r = write(ws, r, ["Note", caveat])

    # ---- Assumptions (the sheet that justifies exporting from here) ----
    ws = sheet("Assumptions")
    widths(ws, 18, 26, 14, 12, 14, 52, 22, 22, 40)
    header(ws, 1, ["Block", "Assumption", "Change (%)", "Locked value", "Unit",
                   "Source URL", "Retrieved (UTC)", "Locked (UTC)", "Rationale"])
    r = first_assumption_row = 2
    by_driver = {str(d.get("driver_id")): d for d in pack.get("drivers") or []}
    spec = pack.get("assumptions") or {}
    for block in budget.ASSUMPTION_BLOCKS:
        for key, value in (spec.get(block) or {}).items():
            provenance = by_driver.get(str(key), {}) if block in ("drivers", "opex") else {}
            r = write(ws, r, [BLOCK_LABELS.get(block, block),
                              provenance.get("name") or key, _num(value),
                              _num(provenance.get("value")), provenance.get("unit"),
                              provenance.get("source_url"), provenance.get("retrieved_at"),
                              provenance.get("locked_at") or pack.get("locked_at"),
                              provenance.get("rationale")],
                      {3: pct, 4: pct})
    # Every watched driver, including the ones this scenario left alone: a board
    # pack has to show what the budget rests on, not only what was shocked.
    untouched = [d for d in pack.get("drivers") or []
                 if str(d.get("driver_id")) not in (spec.get("drivers") or {})]
    if untouched:
        # A scenario that shocks nothing (the as-locked baseline) writes no rows
        # above, and then a spacer would open the sheet on a blank line.
        if r > first_assumption_row:
            r += 1
        r = write(ws, r, ["Unchanged drivers — carried at their locked value"])
        header(ws, r, ["Block", "Driver", "Change (%)", "Locked value", "Unit",
                       "Source URL", "Retrieved (UTC)", "Locked (UTC)", "Rationale"])
        r += 1
        for d in untouched:
            r = write(ws, r, ["Driver price", d.get("name") or d.get("driver_id"), 0.0,
                              _num(d.get("value")), d.get("unit"), d.get("source_url"),
                              d.get("retrieved_at"), d.get("locked_at") or pack.get("locked_at"),
                              d.get("rationale")], {3: pct, 4: pct})

    # ---- Driver bridge ----
    ws = sheet("Driver bridge")
    widths(ws, 34, 20, 14)
    header(ws, 1, ["Step", "EBITDA effect (€)", "Kind"])
    r = 2
    for label, value, kind in bridge_rows(pack.get("ebitda_bridge") or {}):
        r = write(ws, r, [label, _num(value), kind], {2: money})
    opex = pack.get("opex_bridge") or {}
    if opex.get("by_cost_centre"):
        r += 1
        r = write(ws, r, ["Opex by cost centre"])
        header(ws, r, ["Cost centre", "Driver", "Basis", "Change (%)",
                       "Base (€)", "Effect (€)"])
        r += 1
        for line in opex["by_cost_centre"]:
            r = write(ws, r, [line.get("cost_centre"), line.get("driver_id"),
                              line.get("basis"), _num(line.get("pct")),
                              _num(line.get("base_eur")), _num(line.get("delta_eur"))],
                      {4: pct, 5: money, 6: money})
        r = write(ws, r, ["Weighted opex growth (%)", None, None,
                          _num(opex.get("weighted_pct"))], {4: pct})

    # ---- Version diff ----
    ws = sheet("Version diff")
    widths(ws, 30, 22, 18, 18, 18, 40)
    d = pack.get("diff")
    if not d:
        ws["A1"] = "No prior version to compare against."
    else:
        ws["A1"] = (f"v{(d.get('from') or {}).get('version_no')} → "
                    f"v{(d.get('to') or {}).get('version_no')}")
        ws["A1"].font = title_font
        r = 3
        header(ws, r, ["Assumption", "Block", "From (%)", "To (%)", "Change"])
        r += 1
        for block, rows in (d.get("assumptions") or {}).items():
            for row in rows:
                r = write(ws, r, [row.get("key"), BLOCK_LABELS.get(block, block),
                                  _num(row.get("from_pct")), _num(row.get("to_pct")),
                                  row.get("change")], {3: pct, 4: pct})
        if d.get("locked_values"):
            r += 1
            header(ws, r, ["Driver", "Unit", "From", "To", "Change (%)", "Source URL"])
            r += 1
            for row in d["locked_values"]:
                r = write(ws, r, [row.get("name"), row.get("unit"),
                                  _num(row.get("from_value")), _num(row.get("to_value")),
                                  _num(row.get("delta_pct")), row.get("to_source_url")],
                          {3: pct, 4: pct, 5: pct})
        r += 1
        header(ws, r, ["Step", "EBITDA effect (€)", "Kind"])
        r += 1
        for point in d.get("ebitda_bridge") or []:
            r = write(ws, r, [point.get("label"), _num(point.get("value")),
                              point.get("kind")], {2: money})

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def bridge_rows(bridge: dict) -> list[tuple]:
    """The EBITDA bridge as (label, value, kind) rows — every driver on its own
    line. Deliberately unlike `tools._ebitda_bridge_points`, which folds the
    small drivers into "Other drivers" to fit a chart: a spreadsheet has no such
    limit, and a board pack that hides a line is worse than a long one."""
    if not bridge:
        return []
    rows = [("Baseline EBITDA", bridge.get("baseline_ebitda_eur"), "absolute"),
            ("Volume", bridge.get("volume_eur"), "delta"),
            ("Price", bridge.get("price_eur"), "delta")]
    for driver_id, value in sorted((bridge.get("drivers_eur") or {}).items(),
                                   key=lambda kv: -abs(float(kv[1] or 0.0))):
        rows.append((driver_id, value, "delta"))
    rows.append(("Opex", bridge.get("opex_eur"), "delta"))
    rows.append(("Scenario EBITDA", bridge.get("projected_ebitda_eur"), "absolute"))
    return [(label, _num(value), kind) for label, value, kind in rows]


# --------------------------------------------------------------------------
# The board pack
# --------------------------------------------------------------------------

def board_pack_markdown(pack: dict) -> str:
    """The same content as a document you can paste into a board deck.

    Pure text, no dependency — so this export survives a missing openpyxl, and
    the agent can also read it back into a session.
    """
    totals = pack.get("totals") or {}
    out: list[str] = []
    title = pack.get("title") or "Budget"
    out.append(f"# {pack.get('company') or 'Budget'} — {title}")
    out.append("")
    meta = " · ".join(f"**{label}:** {value}" for label, value in (pack.get("meta") or []))
    if meta:
        out += [meta, ""]
    if pack.get("note"):
        out += [pack["note"], ""]
    if pack.get("narrative"):
        out += [pack["narrative"], ""]

    out += ["## Headline", "", "| Metric | Value |", "| --- | ---: |"]
    for key, label in SUMMARY_ROWS:
        out.append(f"| {label} | {_fmt(totals.get(key))} |")
    out.append("")

    spec = pack.get("assumptions") or {}
    by_driver = {str(d.get("driver_id")): d for d in pack.get("drivers") or []}
    if any(spec.get(b) for b in budget.ASSUMPTION_BLOCKS):
        out += ["## Assumptions", "",
                "| Assumption | Type | Change | Locked value | Source | Retrieved |",
                "| --- | --- | ---: | ---: | --- | --- |"]
        for block in budget.ASSUMPTION_BLOCKS:
            for key, value in (spec.get(block) or {}).items():
                p = by_driver.get(str(key), {}) if block in ("drivers", "opex") else {}
                name = p.get("name") or key
                unit = f" {p['unit']}" if p.get("unit") else ""
                source = f"[source]({p['source_url']})" if p.get("source_url") else "—"
                out.append(f"| {name} | {BLOCK_LABELS.get(block, block)} | "
                           f"{_fmt(value, 2)}% | {_fmt(p.get('value'), 2)}{unit} | "
                           f"{source} | {p.get('retrieved_at') or '—'} |")
        out += ["", "_Every driver figure above was recorded from a page the agent "
                    "fetched; the source and retrieval date travel with it._", ""]

    rows = bridge_rows(pack.get("ebitda_bridge") or {})
    if rows:
        out += ["## EBITDA bridge", "", "| Step | € |", "| --- | ---: |"]
        out += [f"| {label} | {_fmt(value)} |" for label, value, _ in rows]
        out.append("")

    opex = pack.get("opex_bridge") or {}
    if opex.get("by_cost_centre"):
        out += ["## Opex by cost centre", "",
                "| Cost centre | Driver | Basis | Change | € |",
                "| --- | --- | --- | ---: | ---: |"]
        for line in opex["by_cost_centre"]:
            out.append(f"| {line.get('cost_centre')} | {line.get('driver_id') or '—'} | "
                       f"{line.get('basis')} | {_fmt(line.get('pct'), 2)}% | "
                       f"{_fmt(line.get('delta_eur'))} |")
        out += ["", f"Weighted opex growth: **{_fmt(opex.get('weighted_pct'), 2)}%**", ""]

    cash = pack.get("cash")
    if cash:
        ct = cash.get("totals") or {}
        cd = cash.get("days") or {}
        trough = cash.get("trough") or {}
        out += ["## Cash", "",
                f"Free cash flow **{_fmt(ct.get('free_cash_flow_eur'))} €** on EBITDA of "
                f"{_fmt(ct.get('ebitda_eur'))} €"
                + (f" ({_fmt(ct.get('cash_conversion_pct'), 1)}% converts)"
                   if ct.get("cash_conversion_pct") is not None else "")
                + f". Cash is lowest in **{trough.get('month') or '—'}** at "
                  f"{_fmt(trough.get('closing_cash_eur'))} €.", ""]
        if cash.get("months_below_minimum"):
            out += [f"⚠ Below the {_fmt(cash.get('min_cash_eur'))} € minimum in "
                    f"{', '.join(cash['months_below_minimum'])}.", ""]
        out += ["| Step | € |", "| --- | ---: |",
                f"| EBITDA | {_fmt(ct.get('ebitda_eur'))} |",
                f"| Working capital | {_fmt(-_num(ct.get('working_capital_change_eur') or 0))} |",
                f"| Cash tax | {_fmt(-_num(ct.get('tax_eur') or 0))} |",
                f"| Capex | {_fmt(-_num(ct.get('capex_eur') or 0))} |",
                f"| Free cash flow | {_fmt(ct.get('free_cash_flow_eur'))} |", "",
                f"Working-capital days: DSO {_fmt(cd.get('dso_days'), 1)}, "
                f"DIO {_fmt(cd.get('dio_days'), 1)}, DPO {_fmt(cd.get('dpo_days'), 1)} "
                f"— cash conversion cycle {_fmt(cd.get('cash_conversion_cycle_days'), 1)} days"
                + (f" ({cd['basis']})." if cd.get("basis") else "."), ""]
        # Every caveat, every time. This is the section a board is most likely to
        # read as a bank balance, and it is not one.
        out += [f"_{caveat}_" for caveat in cash.get("caveats") or []]
        out.append("")

    if pack.get("by_month"):
        out += ["## Monthly P&L", "",
                "| Month | Revenue € | COGS € | Opex € | EBITDA € |",
                "| --- | ---: | ---: | ---: | ---: |"]
        for row in pack["by_month"]:
            out.append(f"| {row.get('month')} | {_fmt(row.get('revenue_eur'))} | "
                       f"{_fmt(row.get('cogs_eur'))} | {_fmt(row.get('opex_eur'))} | "
                       f"{_fmt(row.get('ebitda_eur'))} |")
        out.append("")

    d = pack.get("diff")
    if d:
        a, b = d.get("from") or {}, d.get("to") or {}
        out += [f"## Change from v{a.get('version_no')} to v{b.get('version_no')}", "",
                f"EBITDA moves **{_fmt(d.get('ebitda_delta_eur'))} €**.", ""]
        changes = [(block, row) for block, rows_ in (d.get("assumptions") or {}).items()
                   for row in rows_]
        if changes:
            out += ["| Assumption | Type | From | To |", "| --- | --- | ---: | ---: |"]
            for block, row in changes:
                out.append(f"| {row.get('key')} | {BLOCK_LABELS.get(block, block)} | "
                           f"{_fmt(row.get('from_pct'), 2)} | {_fmt(row.get('to_pct'), 2)} |")
            out.append("")
        if d.get("locked_values"):
            out += ["| Driver re-locked | From | To | Change | Source |",
                    "| --- | ---: | ---: | ---: | --- |"]
            for row in d["locked_values"]:
                source = (f"[source]({row['to_source_url']})"
                          if row.get("to_source_url") else "—")
                out.append(f"| {row.get('name')} | {_fmt(row.get('from_value'), 2)} | "
                           f"{_fmt(row.get('to_value'), 2)} | "
                           f"{_fmt(row.get('delta_pct'), 2)}% | {source} |")
            out.append("")
        if d.get("ebitda_bridge"):
            out += ["| Step | € |", "| --- | ---: |"]
            out += [f"| {p.get('label')} | {_fmt(p.get('value'))} |"
                    for p in d["ebitda_bridge"]]
            out.append("")

    out.append(f"_Exported {pack.get('generated_at')} · "
               f"figures computed by the budgeting agent, not typed._")
    return "\n".join(out)


# --------------------------------------------------------------------------
# Formatting helpers
# --------------------------------------------------------------------------

def _num(value):
    """A float where one is meant, None where the figure is genuinely absent —
    an em dash in a numeric spreadsheet column is worse than an empty cell."""
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value if isinstance(value, str) else None


def _fmt(value, digits: int = 0) -> str:
    n = _num(value)
    if n is None:
        return "—"
    if isinstance(n, str):
        return n
    return f"{n:,.{digits}f}"


def _stamp(epoch) -> str | None:
    try:
        return datetime.fromtimestamp(float(epoch), tz=timezone.utc).isoformat(
            timespec="seconds")
    except (TypeError, ValueError):
        return None
